from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import base64
import asyncio
import re
import random
import string
import resend
from openpyxl import load_workbook, Workbook
from io import BytesIO
import google.generativeai as genai


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Banco master — armazena todas as federações
db_master = client['squashrank_master']

def get_federation_db(slug: str):
    return client[f"squashrank_{slug}"]

# JWT Configuration
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production-123456789')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Email Configuration
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'gustavopizatto@hotmail.com')
OWNER_EMAIL = ADMIN_EMAIL  # dono do sistema

if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# ============= HELPER FUNCTIONS =============

def calculate_set_result(score: List[str]) -> str:
    """
    Calculate set result from score list.
    Example: ["11-2", "11-9", "9-11", "11-7"] -> "3-1"
    """
    if not score:
        return "0-0"
    
    player1_sets = 0
    player2_sets = 0
    
    for game in score:
        try:
            parts = game.split('-')
            if len(parts) == 2:
                score1 = int(parts[0])
                score2 = int(parts[1])
                if score1 > score2:
                    player1_sets += 1
                else:
                    player2_sets += 1
        except (ValueError, IndexError):
            continue
    
    return f"{player1_sets}-{player2_sets}"

def calculate_placement_from_round(round_name: str, is_winner: bool) -> int:
    """
    Calculate tournament placement based on round and whether player won.
    
    Rules:
    - Final: Winner = 1st, Loser = 2nd
    - Semi Final: Loser = 3rd/4th
    - Quarter Final: Loser = 5th-8th
    - Round of 16: Loser = 9th-16th
    - Round of 32: Loser = 17th-32nd
    """
    round_lower = round_name.lower()
    
    if 'final' in round_lower and 'semi' not in round_lower and 'quarter' not in round_lower:
        # Final
        return 1 if is_winner else 2
    elif 'semi' in round_lower:
        # Semi Final
        return 1 if is_winner else 3  # Winner goes to final, loser is 3rd/4th
    elif 'quarter' in round_lower:
        # Quarter Final
        return 3 if is_winner else 5  # Winner goes to semi, loser is 5th-8th
    elif '16' in round_lower:
        # Round of 16
        return 5 if is_winner else 9
    elif '32' in round_lower:
        # Round of 32
        return 9 if is_winner else 17
    else:
        # Group stage or other
        return 9  # Default placement

def get_result_label(placement: int) -> str:
    """Get human-readable result label from placement"""
    labels = {
        1: "Champion",
        2: "Runner-up",
        3: "Semi Finalist",
        4: "Semi Finalist",
        5: "Quarter Finalist",
        6: "Quarter Finalist",
        7: "Quarter Finalist",
        8: "Quarter Finalist"
    }
    return labels.get(placement, f"{placement}º Place")

# ============= SUBSCRIPTION PLANS =============

SUBSCRIPTION_PLANS = {
    "mensal": {
        "name": "Plano Mensal",
        "price": 99.90,  # R$ 99,90/mês
        "currency": "brl",
        "duration_days": 30
    },
    "anual": {
        "name": "Plano Anual",
        "price": 999.00,  # R$ 999,00/ano (economiza ~17%)
        "currency": "brl",
        "duration_days": 365
    }
}

def check_subscription_active(subscription: Dict) -> bool:
    """Check if subscription is active and not expired"""
    if not subscription:
        return False
    
    if subscription['status'] not in ['active']:
        return False
    
    # Check if not expired
    end_date = subscription['end_date']
    if isinstance(end_date, str):
        end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    
    return datetime.now(timezone.utc) < end_date

# ============= MODELS =============

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    hashed_password: str
    is_admin: bool = True
    is_owner: bool = False       # dono do sistema — nunca bloqueado
    is_approved: bool = False    # aprovado manualmente pelo dono após pagamento
    federation_name: Optional[str] = None
    subscription_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    password: str

class FederationCreate(BaseModel):
    federation_name: str
    email: str
    password: str
    plan_type: str  # "mensal" or "anual"
    start_trial: bool = False

class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str

class Subscription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    plan_type: str  # "mensal" or "anual"
    status: str = "active"  # active, inactive, canceled, expired
    is_trial: bool = False
    start_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    end_date: datetime
    next_billing_date: Optional[datetime] = None
    stripe_subscription_id: Optional[str] = None
    stripe_customer_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    session_id: str
    amount: float
    currency: str
    payment_status: str = "pending"  # pending, paid, failed, expired
    metadata: Optional[Dict[str, str]] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RegistrationRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password_hash: str
    approval_code: str
    status: str = "pending"  # pending, approved, rejected
    requested_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    expires_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc) + timedelta(hours=24))

class RegistrationRequestCreate(BaseModel):
    username: str
    password: str

class RegistrationApproval(BaseModel):
    username: str
    approval_code: str

class Player(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    photo_url: Optional[str] = None
    city: Optional[str] = None
    academy: Optional[str] = None
    coach: Optional[str] = None
    main_class: Optional[str] = None  # Classe principal que joga
    birth_date: Optional[str] = None  # ISO date string 'YYYY-MM-DD'
    gender: Optional[str] = None      # 'Masculina' ou 'Feminina'
    is_federated: bool = True         # Se False, não recebe pontos no ranking
    email: Optional[str] = None
    phone: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PlayerCreate(BaseModel):
    name: str
    photo_url: Optional[str] = None
    city: Optional[str] = None
    academy: Optional[str] = None
    coach: Optional[str] = None
    main_class: Optional[str] = None
    birth_date: Optional[str] = None
    gender: Optional[str] = None
    is_federated: bool = True
    email: Optional[str] = None
    phone: Optional[str] = None

class Tournament(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    date: datetime
    location: Optional[str] = None
    is_completed: bool = False
    bracket_link: Optional[str] = None
    photos_link: Optional[str] = None
    stream_link: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TournamentCreate(BaseModel):
    name: str
    date: datetime
    location: Optional[str] = None
    is_completed: bool = False
    bracket_link: Optional[str] = None
    photos_link: Optional[str] = None
    stream_link: Optional[str] = None
    
class Result(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tournament_id: str
    player_id: str
    player_name: str  # Denormalized for performance
    class_category: str  # "1ª", "2ª", "3ª", "4ª", "5ª", "6ª", "Duplas"
    gender_category: str  # "Masculina", "Feminina"
    placement: int
    points: float = 0.0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ResultCreate(BaseModel):
    tournament_id: str
    player_id: str
    class_category: str
    gender_category: str
    placement: int

class RankingConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    formula: str = "top_n"  # "sum_all", "top_n", "decay"
    top_n_count: int = 5
    points_table: Dict[str, float] = Field(default_factory=lambda: {
        "1": 100, "2": 75, "3": 50, "4": 50, "5": 25, "6": 25, "7": 25, "8": 25,
        "9": 10, "10": 10, "11": 10, "12": 10, "13": 10, "14": 10, "15": 10, "16": 10
    })
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RankingConfigUpdate(BaseModel):
    formula: str
    top_n_count: int
    points_table: Dict[str, float]

class ImportResult(BaseModel):
    player_name: str
    placement: int
    category: str

class ImportData(BaseModel):
    tournament_name: str
    results: List[ImportResult]

class Match(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tournament_id: str  # Obrigatório - ID do torneio
    tournament_name: str
    category: str  # "1ª", "2ª", etc.
    player1_id: str
    player1_name: str
    player2_id: str
    player2_name: str
    winner_id: str
    score: List[str]  # ["11-7", "8-11", "11-6", "11-9"]
    round: str  # "Final", "Semi Final", "Quarter Final", "Round of 16", etc.
    date: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MatchCreate(BaseModel):
    tournament_id: str
    category: str
    player1_id: str
    player2_id: str
    winner_id: str
    score: List[str]
    round: str
    date: datetime
class FederationPublic(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    slug: str
    logo_url: Optional[str] = None
    state: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FederationCreate(BaseModel):
    name: str
    slug: str
    logo_url: Optional[str] = None
    state: Optional[str] = None

# ============= AUTH FUNCTIONS =============

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def generate_approval_code() -> str:
    """Generate a 6-character alphanumeric code"""
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

async def send_approval_email(username: str, approval_code: str):
    """Send approval code to admin email"""
    if not RESEND_API_KEY:
        logger.warning("RESEND_API_KEY not set. Skipping email.")
        return
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
            .content {{ background: #f8f9fa; padding: 30px; border-radius: 0 0 10px 10px; }}
            .code-box {{ background: #22c55e; color: white; font-size: 32px; font-weight: bold; padding: 20px; text-align: center; border-radius: 8px; margin: 20px 0; letter-spacing: 4px; }}
            .info {{ background: white; padding: 15px; border-left: 4px solid #3b82f6; margin: 20px 0; }}
            .footer {{ text-align: center; color: #666; font-size: 12px; margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🎾 Nova Solicitação de Registro</h1>
                <p>SquashRank Pro - Federação de Squash do Paraná</p>
            </div>
            <div class="content">
                <h2>Tentativa de Registro Detectada</h2>
                <div class="info">
                    <strong>Usuário:</strong> {username}<br>
                    <strong>Data/Hora:</strong> {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M:%S')} UTC
                </div>
                <p>Alguém está tentando se registrar como administrador no sistema SquashRank Pro.</p>
                <p><strong>Código de Aprovação:</strong></p>
                <div class="code-box">{approval_code}</div>
                <p>Para aprovar este registro:</p>
                <ol>
                    <li>Verifique se você reconhece este usuário</li>
                    <li>Forneça o código acima para o usuário</li>
                    <li>O usuário deve inserir este código para completar o registro</li>
                </ol>
                <p><strong>⏰ Este código expira em 24 horas.</strong></p>
                <div class="footer">
                    <p>Este é um email automático do sistema SquashRank Pro</p>
                    <p>Se você não esperava este email, ignore-o com segurança</p>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    
    params = {
        "from": SENDER_EMAIL,
        "to": [ADMIN_EMAIL],
        "subject": f"🎾 Nova Solicitação de Registro - {username}",
        "html": html_content
    }
    
    try:
        await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Approval email sent to {ADMIN_EMAIL} for user {username}")
    except Exception as e:
        logger.error(f"Failed to send approval email: {str(e)}")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"username": username}, {"_id": 0})
    if user is None:
        raise credentials_exception
    return User(**user)

# ============= ROUTES =============

@api_router.get("/")
async def root():
    return {"message": "SquashRank Pro API"}

@api_router.get("/federations")
async def list_federations():
    feds = await db_master.federations.find({"is_active": True}, {"_id": 0}).to_list(1000)
    for fed in feds:
        fed['total_players'] = await db.players.count_documents({})
        fed['total_tournaments'] = await db.tournaments.count_documents({})
    return feds

@api_router.get("/federations/{slug}")
async def get_federation(slug: str):
    fed = await db_master.federations.find_one({"slug": slug}, {"_id": 0})
    if not fed:
        raise HTTPException(status_code=404, detail="Federation not found")
    return fed

@api_router.post("/federations")
async def create_federation(data: FederationCreate):
    existing = await db_master.federations.find_one({"slug": data.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Slug already in use")
    fed = FederationPublic(**data.model_dump())
    doc = fed.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db_master.federations.insert_one(doc)
    return fed

# Auth Routes
@api_router.post("/auth/request-registration")
async def request_registration(user_data: RegistrationRequestCreate):
    # Check if username already exists
    existing_user = await db.users.find_one({"username": user_data.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if there's already a pending request
    existing_request = await db.registration_requests.find_one({
        "username": user_data.username,
        "status": "pending"
    })
    
    if existing_request:
        raise HTTPException(
            status_code=400, 
            detail="Registration request already pending. Check your email for the approval code."
        )
    
    # Generate approval code
    approval_code = generate_approval_code()
    
    # Hash password
    hashed_password = get_password_hash(user_data.password)
    
    # Create registration request
    request = RegistrationRequest(
        username=user_data.username,
        password_hash=hashed_password,
        approval_code=approval_code
    )
    
    doc = request.model_dump()
    doc['requested_at'] = doc['requested_at'].isoformat()
    doc['expires_at'] = doc['expires_at'].isoformat()
    await db.registration_requests.insert_one(doc)
    
    # Send approval email to admin
    await send_approval_email(user_data.username, approval_code)
    
    return {
        "message": "Registration request submitted. Please wait for admin approval.",
        "username": user_data.username
    }

@api_router.post("/auth/complete-registration", response_model=Token)
async def complete_registration(approval_data: RegistrationApproval):
    # Find registration request
    request_doc = await db.registration_requests.find_one({
        "username": approval_data.username,
        "status": "pending"
    }, {"_id": 0})
    
    if not request_doc:
        raise HTTPException(status_code=404, detail="Registration request not found or already processed")
    
    # Check if expired
    expires_at = datetime.fromisoformat(request_doc['expires_at'])
    if datetime.now(timezone.utc) > expires_at:
        await db.registration_requests.update_one(
            {"id": request_doc['id']},
            {"$set": {"status": "expired"}}
        )
        raise HTTPException(status_code=400, detail="Registration request expired. Please request again.")
    
    # Verify approval code
    if request_doc['approval_code'] != approval_data.approval_code:
        raise HTTPException(status_code=401, detail="Invalid approval code")
    
    # Create user
    user = User(
        username=request_doc['username'],
        hashed_password=request_doc['password_hash'],
        is_admin=True
    )
    
    user_doc = user.model_dump()
    user_doc['created_at'] = user_doc['created_at'].isoformat()
    await db.users.insert_one(user_doc)
    
    # Mark request as approved
    await db.registration_requests.update_one(
        {"id": request_doc['id']},
        {"$set": {"status": "approved"}}
    )
    
    # Generate token
    access_token = create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    user_doc = await db.users.find_one({"username": user_data.username}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    user = User(**user_doc)
    if not verify_password(user_data.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    access_token = create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserLogin):
    existing_user = await db.users.find_one({"username": user_data.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists")

    hashed_password = get_password_hash(user_data.password)

    # O dono do sistema (ADMIN_EMAIL) é aprovado automaticamente
    is_owner = (user_data.username.lower() == ADMIN_EMAIL.lower())

    user = User(
        username=user_data.username,
        hashed_password=hashed_password,
        is_admin=True,
        is_owner=is_owner,
        is_approved=is_owner  # dono já aprovado, outros aguardam
    )

    user_doc = user.model_dump()
    user_doc['created_at'] = user_doc['created_at'].isoformat()
    await db.users.insert_one(user_doc)

    access_token = create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}
@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    return {
        "username": current_user.username,
        "is_admin": current_user.is_admin,
        "is_owner": current_user.is_owner,
        "is_approved": current_user.is_approved or current_user.is_owner
    }

@api_router.get("/auth/approval-status")
async def get_approval_status(current_user: User = Depends(get_current_user)):
    """Verifica se o usuário está aprovado para usar o sistema"""
    approved = current_user.is_approved or current_user.is_owner
    return {"is_approved": approved, "is_owner": current_user.is_owner}

@api_router.get("/admin/pending-users")
async def get_pending_users(current_user: User = Depends(get_current_user)):
    """Lista usuários aguardando aprovação — apenas para o dono"""
    if not current_user.is_owner:
        raise HTTPException(status_code=403, detail="Apenas o dono pode ver usuários pendentes")
    pending = await db.users.find(
        {"is_approved": False, "is_owner": {"$ne": True}},
        {"_id": 0, "hashed_password": 0}
    ).to_list(100)
    return pending

@api_router.post("/admin/approve-user/{username}")
async def approve_user(username: str, current_user: User = Depends(get_current_user)):
    """Aprova um usuário manualmente — apenas para o dono"""
    if not current_user.is_owner:
        raise HTTPException(status_code=403, detail="Apenas o dono pode aprovar usuários")
    result = await db.users.update_one(
        {"username": username},
        {"$set": {"is_approved": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return {"message": f"Usuário {username} aprovado com sucesso"}

@api_router.get("/admin/all-users")
async def get_all_users(current_user: User = Depends(get_current_user)):
    """Lista todos os usuários — apenas para o dono"""
    if not current_user.is_owner:
        raise HTTPException(status_code=403, detail="Apenas o dono pode ver todos os usuários")
    users = await db.users.find({}, {"_id": 0, "hashed_password": 0}).to_list(200)
    return users

@api_router.post("/admin/revoke-user/{username}")
async def revoke_user(username: str, current_user: User = Depends(get_current_user)):
    """Revoga acesso de um usuário — apenas para o dono"""
    if not current_user.is_owner:
        raise HTTPException(status_code=403, detail="Apenas o dono pode revogar usuários")
    result = await db.users.update_one(
        {"username": username},
        {"$set": {"is_approved": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return {"message": f"Acesso de {username} revogado"}

# Player Routes
@api_router.get("/players", response_model=List[Player])
async def get_players():
    players = await db.players.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    for player in players:
        if isinstance(player.get('created_at'), str):
            player['created_at'] = datetime.fromisoformat(player['created_at'])
    return players

@api_router.post("/players", response_model=Player)
async def create_player(player_data: PlayerCreate, current_user: User = Depends(get_current_user)):
    player = Player(**player_data.model_dump())
    doc = player.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.players.insert_one(doc)
    return player

@api_router.put("/players/{player_id}", response_model=Player)
async def update_player(player_id: str, player_data: PlayerCreate, current_user: User = Depends(get_current_user)):
    updated_data = player_data.model_dump()
    result = await db.players.update_one({"id": player_id}, {"$set": updated_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Player not found")
    
    player_doc = await db.players.find_one({"id": player_id}, {"_id": 0})
    if isinstance(player_doc.get('created_at'), str):
        player_doc['created_at'] = datetime.fromisoformat(player_doc['created_at'])
    return Player(**player_doc)

@api_router.delete("/players/{player_id}")
async def delete_player(player_id: str, current_user: User = Depends(get_current_user)):
    result = await db.players.delete_one({"id": player_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Player not found")
    return {"message": "Player deleted"}

@api_router.post("/players/merge")
async def merge_players(
    keep_id: str,
    remove_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Mescla dois jogadores duplicados.
    - keep_id: jogador que será mantido (nome correto/completo)
    - remove_id: jogador duplicado que será removido
    Todos os resultados e partidas do duplicado são transferidos para o jogador mantido.
    """
    keep = await db.players.find_one({"id": keep_id}, {"_id": 0})
    remove = await db.players.find_one({"id": remove_id}, {"_id": 0})

    if not keep:
        raise HTTPException(status_code=404, detail="Jogador principal não encontrado")
    if not remove:
        raise HTTPException(status_code=404, detail="Jogador duplicado não encontrado")
    if keep_id == remove_id:
        raise HTTPException(status_code=400, detail="Os jogadores devem ser diferentes")

    # Copia campos não vazios do duplicado para o jogador mantido
    fields_to_copy = ['photo_url', 'city', 'academy', 'coach', 'main_class',
                      'birth_date', 'gender', 'email', 'phone']
    update_fields = {}
    for field in fields_to_copy:
        keep_val = keep.get(field)
        remove_val = remove.get(field)
        # Copia do duplicado se o campo principal estiver vazio/None
        if not keep_val and remove_val:
            update_fields[field] = remove_val

    if update_fields:
        await db.players.update_one({"id": keep_id}, {"$set": update_fields})
        keep.update(update_fields)  # atualiza local para usar o nome correto abaixo

    results_updated = 0
    matches_updated = 0

    # Transfere resultados — mas evita duplicar resultado no mesmo torneio/classe
    remove_results = await db.results.find({"player_id": remove_id}, {"_id": 0}).to_list(1000)
    for result in remove_results:
        # Checa se já existe resultado igual para o jogador mantido
        existing = await db.results.find_one({
            "player_id": keep_id,
            "tournament_id": result["tournament_id"],
            "class_category": result["class_category"],
            "gender_category": result["gender_category"]
        })
        if existing:
            # Mantém o melhor placement
            if result["placement"] < existing["placement"]:
                await db.results.update_one(
                    {"id": existing["id"]},
                    {"$set": {"placement": result["placement"], "points": result["points"]}}
                )
            await db.results.delete_one({"id": result["id"]})
        else:
            await db.results.update_one(
                {"id": result["id"]},
                {"$set": {"player_id": keep_id, "player_name": keep["name"]}}
            )
            results_updated += 1

    # Transfere partidas
    matches_p1 = await db.matches.find({"player1_id": remove_id}, {"_id": 0}).to_list(1000)
    for match in matches_p1:
        await db.matches.update_one(
            {"id": match["id"]},
            {"$set": {"player1_id": keep_id, "player1_name": keep["name"],
                      "winner_id": keep_id if match["winner_id"] == remove_id else match["winner_id"]}}
        )
        matches_updated += 1

    matches_p2 = await db.matches.find({"player2_id": remove_id}, {"_id": 0}).to_list(1000)
    for match in matches_p2:
        await db.matches.update_one(
            {"id": match["id"]},
            {"$set": {"player2_id": keep_id, "player2_name": keep["name"],
                      "winner_id": keep_id if match["winner_id"] == remove_id else match["winner_id"]}}
        )
        matches_updated += 1

    # Remove o jogador duplicado
    await db.players.delete_one({"id": remove_id})

    return {
        "message": f"Jogadores mesclados com sucesso",
        "kept_player": keep["name"],
        "removed_player": remove["name"],
        "results_transferred": results_updated,
        "matches_transferred": matches_updated
    }


@api_router.post("/players/upload-photo")
async def upload_photo(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Arquivo deve ser uma imagem")
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Imagem deve ter no maximo 5MB")
    base64_encoded = base64.b64encode(contents).decode('utf-8')
    photo_url = f"data:{file.content_type};base64,{base64_encoded}"
    return {"photo_url": photo_url}

@api_router.get("/players/template")
async def download_players_template():
    """Generate and download Excel template for players import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Players"
    
    # Headers
    headers = ["Name", "City", "Academy", "Coach", "Main Class"]
    ws.append(headers)
    
    # Example row
    example = ["João Silva", "Curitiba", "Academia XYZ", "Pedro Treinador", "1ª"]
    ws.append(example)
    
    # Style headers (bold)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_jogadores.xlsx"}
    )

@api_router.post("/import-players-excel")
async def import_players_excel(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """
    Importa jogadores do arquivo Excel de Filiação (Google Forms / Tournament Planner).
    Formato esperado (aba 'Respostas ao formulário 1' ou a primeira aba):
    Col A: Carimbo de data/hora
    Col B: E-mail
    Col C: Nome Completo
    Col D: Gênero
    Col E: Data de nascimento
    Col F: CPF
    Col G: Celular
    Col H: Cidade
    Col I: Estado
    Col J: Clube/Academia
    Col K: Treinador
    (colunas L+ ignoradas)
    """
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")

    try:
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))

        # Tenta usar a primeira aba que contém dados de filiação
        sheet = workbook.active

        players_created = 0
        players_updated = 0
        errors = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[2]:  # Nome Completo é coluna C (índice 2)
                continue

            try:
                name = str(row[2]).strip()
                if not name:
                    continue

                gender   = normalize_gender(str(row[3]).strip()) if row[3] else None
                email    = str(row[1]).strip() if row[1] else None
                city_raw = str(row[7]).strip() if row[7] else None
                state    = str(row[8]).strip() if row[8] else None
                academy  = str(row[9]).strip() if row[9] else None
                coach    = str(row[10]).strip() if row[10] else None
                phone_raw= row[6]
                phone    = str(int(float(phone_raw))) if phone_raw else None

                # Cidade + Estado
                city = None
                if city_raw and state:
                    city = f"{city_raw}, {state}"
                elif city_raw:
                    city = city_raw

                # Data de nascimento
                birth_date = None
                if row[4]:
                    from datetime import date as date_type
                    if isinstance(row[4], (datetime, date_type)):
                        birth_date = row[4].strftime('%Y-%m-%d')
                    else:
                        try:
                            from dateutil import parser as date_parser
                            birth_date = date_parser.parse(str(row[4])).strftime('%Y-%m-%d')
                        except Exception:
                            birth_date = str(row[4])

                # Checa se jogador já existe pelo nome (case-insensitive)
                existing = await db.players.find_one(
                    {"name": {"$regex": f"^{name}$", "$options": "i"}},
                    {"_id": 0}
                )

                update_data = {
                    "gender": gender,
                    "birth_date": birth_date,
                    "email": email,
                    "phone": phone,
                    "city": city,
                    "academy": academy,
                    "coach": coach,
                }

                if existing:
                    await db.players.update_one({"id": existing['id']}, {"$set": update_data})
                    players_updated += 1
                else:
                    player = Player(name=name, **update_data)
                    doc = player.model_dump()
                    doc['created_at'] = doc['created_at'].isoformat()
                    await db.players.insert_one(doc)
                    players_created += 1

            except Exception as e:
                errors.append(f"Linha {row_idx}: {str(e)}")

        return {
            "players_created": players_created,
            "players_updated": players_updated,
            "errors": errors
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar Excel: {str(e)}")

@api_router.get("/players/{player_id}/details")
async def get_player_details(player_id: str):
    # Get player info
    player = await db.players.find_one({"id": player_id}, {"_id": 0})
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    
    # Get player results
    results = await db.results.find({"player_id": player_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Get tournaments for these results
    tournament_ids = list(set([r['tournament_id'] for r in results]))
    tournaments_cursor = db.tournaments.find({"id": {"$in": tournament_ids}}, {"_id": 0})
    tournaments_list = await tournaments_cursor.to_list(100)
    tournaments_dict = {t['id']: t for t in tournaments_list}
    
    # Build recent tournaments with results
    recent_tournaments = []
    for result in results[:10]:  # Last 10 tournaments
        tournament = tournaments_dict.get(result['tournament_id'])
        if tournament:
            recent_tournaments.append({
                "tournament_name": tournament['name'],
                "tournament_date": tournament['date'],
                "class_category": result['class_category'],
                "gender_category": result['gender_category'],
                "placement": result['placement'],
                "points": result['points']
            })
    
    # Get ranking position for each class/category combination
    rankings = {}
    classes_played = list(set([(r['class_category'], r['gender_category']) for r in results]))
    
    config = await get_ranking_config()
    
    for class_cat, gender_cat in classes_played:
        # Get all results for this class/category
        all_results = await db.results.find(
            {"class_category": class_cat, "gender_category": gender_cat},
            {"_id": 0}
        ).to_list(10000)
        
        # Calculate rankings
        player_results = {}
        for r in all_results:
            pid = r['player_id']
            if pid not in player_results:
                player_results[pid] = {'player_name': r['player_name'], 'results': []}
            player_results[pid]['results'].append(r['points'])
        
        # Calculate points based on formula
        ranking_list = []
        for pid, data in player_results.items():
            points_list = sorted(data['results'], reverse=True)
            if config.formula == "sum_all":
                total_points = sum(points_list)
            elif config.formula == "top_n":
                total_points = sum(points_list[:config.top_n_count])
            else:
                total_points = sum([p * (0.9 ** i) for i, p in enumerate(points_list)])
            
            ranking_list.append({
                'player_id': pid,
                'total_points': total_points
            })
        
        ranking_list.sort(key=lambda x: x['total_points'], reverse=True)
        
        # Find player position
        for i, r in enumerate(ranking_list, 1):
            if r['player_id'] == player_id:
                rankings[f"{class_cat}_{gender_cat}"] = {
                    "rank": i,
                    "total": len(ranking_list),
                    "points": round(r['total_points'], 2),
                    "class": class_cat,
                    "category": gender_cat
                }
                break
    
    # Get player matches
    matches = await db.matches.find(
        {"$or": [{"player1_id": player_id}, {"player2_id": player_id}]},
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Process matches to add opponent and result info
    match_history = []
    last_match = None
    head_to_head = {}  # Dictionary to store head-to-head stats
    
    for idx, match in enumerate(matches):
        is_player1 = match['player1_id'] == player_id
        opponent_name = match['player2_name'] if is_player1 else match['player1_name']
        opponent_id = match['player2_id'] if is_player1 else match['player1_id']
        is_winner = match['winner_id'] == player_id
        
        # Calculate set result
        set_result = calculate_set_result(match['score'])
        score_formatted = f"{', '.join(match['score'])} ({set_result})"
        
        match_info = {
            "id": match['id'],
            "tournament_name": match['tournament_name'],
            "opponent_name": opponent_name,
            "opponent_id": opponent_id,
            "result": "Win" if is_winner else "Loss",
            "score": " ".join(match['score']),
            "score_formatted": score_formatted,
            "set_result": set_result,
            "round": match['round'],
            "date": match['date']
        }
        
        match_history.append(match_info)
        
        # Save first match as last_match
        if idx == 0:
            last_match = match_info
        
        # Calculate head-to-head stats
        if opponent_id not in head_to_head:
            head_to_head[opponent_id] = {
                "opponent_id": opponent_id,
                "opponent_name": opponent_name,
                "matches_played": 0,
                "wins": 0,
                "losses": 0
            }
        
        head_to_head[opponent_id]["matches_played"] += 1
        if is_winner:
            head_to_head[opponent_id]["wins"] += 1
        else:
            head_to_head[opponent_id]["losses"] += 1
    
    # Convert head_to_head to sorted list (by matches played)
    head_to_head_list = sorted(
        head_to_head.values(),
        key=lambda x: x["matches_played"],
        reverse=True
    )
    
    return {
        "player": player,
        "recent_tournaments": recent_tournaments,
        "rankings": rankings,
        "total_tournaments": len(results),
        "match_history": match_history,
        "last_match": last_match,
        "head_to_head": head_to_head_list
    }

# Tournament Routes
@api_router.get("/tournaments", response_model=List[Tournament])
async def get_tournaments():
    tournaments = await db.tournaments.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    for tournament in tournaments:
        if isinstance(tournament.get('date'), str):
            tournament['date'] = datetime.fromisoformat(tournament['date'])
        if isinstance(tournament.get('created_at'), str):
            tournament['created_at'] = datetime.fromisoformat(tournament['created_at'])
    return tournaments

@api_router.get("/tournaments/{tournament_id}/results")
async def get_tournament_results(tournament_id: str):
    # Get tournament info
    tournament = await db.tournaments.find_one({"id": tournament_id}, {"_id": 0})
    if not tournament:
        raise HTTPException(status_code=404, detail="Tournament not found")
    
    # Get all results for this tournament
    results = await db.results.find({"tournament_id": tournament_id}, {"_id": 0}).to_list(1000)
    
    # Group by class and category
    grouped_results = {}
    for result in results:
        key = f"{result['class_category']}_{result['gender_category']}"
        if key not in grouped_results:
            grouped_results[key] = {
                "class": result['class_category'],
                "category": result['gender_category'],
                "results": []
            }
        grouped_results[key]["results"].append({
            "player_id": result['player_id'],
            "player_name": result['player_name'],
            "placement": result['placement'],
            "points": result['points']
        })
    
    # Sort results by placement
    for key in grouped_results:
        grouped_results[key]["results"].sort(key=lambda x: x['placement'])
    
    return {
        "tournament": tournament,
        "results": list(grouped_results.values())
    }

@api_router.post("/tournaments", response_model=Tournament)
async def create_tournament(tournament_data: TournamentCreate, current_user: User = Depends(get_current_user)):
    tournament = Tournament(**tournament_data.model_dump())
    doc = tournament.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.tournaments.insert_one(doc)
    return tournament

@api_router.put("/tournaments/{tournament_id}", response_model=Tournament)
async def update_tournament(tournament_id: str, tournament_data: TournamentCreate, current_user: User = Depends(get_current_user)):
    updated_data = tournament_data.model_dump()
    updated_data['date'] = updated_data['date'].isoformat()
    result = await db.tournaments.update_one({"id": tournament_id}, {"$set": updated_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tournament not found")
    
    tournament_doc = await db.tournaments.find_one({"id": tournament_id}, {"_id": 0})
    if isinstance(tournament_doc.get('date'), str):
        tournament_doc['date'] = datetime.fromisoformat(tournament_doc['date'])
    if isinstance(tournament_doc.get('created_at'), str):
        tournament_doc['created_at'] = datetime.fromisoformat(tournament_doc['created_at'])
    return Tournament(**tournament_doc)

@api_router.delete("/tournaments/{tournament_id}")
async def delete_tournament(tournament_id: str, current_user: User = Depends(get_current_user)):
    result = await db.tournaments.delete_one({"id": tournament_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tournament not found")
    return {"message": "Tournament deleted"}

@api_router.get("/tournaments/{tournament_id}/matches")
async def get_tournament_matches(tournament_id: str, category: Optional[str] = None):
    # Get tournament info
    tournament = await db.tournaments.find_one({"id": tournament_id}, {"_id": 0})
    if not tournament:
        raise HTTPException(status_code=404, detail="Tournament not found")
    
    # Build query
    query = {"tournament_id": tournament_id}
    if category:
        query["category"] = category
    
    # Get matches
    matches = await db.matches.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    for match in matches:
        if isinstance(match.get('date'), str):
            match['date'] = datetime.fromisoformat(match['date'])
        if isinstance(match.get('created_at'), str):
            match['created_at'] = datetime.fromisoformat(match['created_at'])
    
    # Group by category
    grouped_matches = {}
    for match in matches:
        cat = match['category']
        if cat not in grouped_matches:
            grouped_matches[cat] = []
        grouped_matches[cat].append(match)
    
    return {
        "tournament": tournament,
        "categories": list(grouped_matches.keys()),
        "matches": grouped_matches
    }

# Result Routes
@api_router.get("/results", response_model=List[Result])
async def get_results():
    results = await db.results.find({}, {"_id": 0}).to_list(10000)
    for result in results:
        if isinstance(result.get('created_at'), str):
            result['created_at'] = datetime.fromisoformat(result['created_at'])
    return results

@api_router.post("/results", response_model=Result)
async def create_result(result_data: ResultCreate, current_user: User = Depends(get_current_user)):
    # Get player name
    player = await db.players.find_one({"id": result_data.player_id}, {"_id": 0})
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    
    # Get points from config
    config = await get_ranking_config()
    points = config.points_table.get(str(result_data.placement), 0.0)
    
    result = Result(
        **result_data.model_dump(),
        player_name=player['name'],
        points=points
    )
    doc = result.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.results.insert_one(doc)
    return result

@api_router.delete("/results/{result_id}")
async def delete_result(result_id: str, current_user: User = Depends(get_current_user)):
    result = await db.results.delete_one({"id": result_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Result not found")
    return {"message": "Result deleted"}

@api_router.get("/results/template")
async def download_results_template():
    """Generate and download Excel template for results import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Headers
    headers = ["Player", "Classe", "Position"]
    ws.append(headers)
    
    # Example row
    example = ["João Silva", "Masculina A", "Champion"]
    ws.append(example)
    
    # Style headers (bold)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_resultados.xlsx"}
    )


def normalize_gender(value: str) -> str:
    """
    Normaliza qualquer variação de gênero para o padrão do sistema.
    Masculino/masculino -> Masculina
    Feminino/feminino  -> Feminina
    Misto/misto        -> Mista
    """
    if not value:
        return value
    v = value.strip().lower()
    if v in ('masculino', 'masculina', 'm', 'masc'):
        return 'Masculina'
    if v in ('feminino', 'feminina', 'f', 'fem'):
        return 'Feminina'
    if v in ('misto', 'mista', 'duplas'):
        return 'Mista'
    return value.strip()


async def get_or_create_player(name: str, players_dict: dict, gender: str = None) -> dict:
    """
    Busca jogador pelo nome (case-insensitive).
    Se não existir, cria automaticamente e atualiza players_dict.
    """
    key = name.lower().strip()
    if key in players_dict:
        return players_dict[key]

    # Tenta busca direta no banco (caso players_dict esteja desatualizado)
    existing = await db.players.find_one(
        {"name": {"$regex": f"^{re.escape(name.strip())}$", "$options": "i"}},
        {"_id": 0}
    )
    if existing:
        players_dict[key] = existing
        return existing

    # Cria novo jogador
    new_player = Player(name=name.strip(), gender=gender)
    doc = new_player.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.players.insert_one(doc)
    players_dict[key] = doc
    return doc


@api_router.post("/results/import")
async def import_results(
    tournament_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Importa resultados de arquivo Excel do Tournament Planner.
    Formato: múltiplas abas, cada aba = uma categoria.
    Nome da aba: ex. '1ª Classe Masculina - 1ª Classe', 'Duplas - Duplas'
    Colunas: Member ID1 | Member ID2 | Player1 | Player2 | Position | Position2
    - Para individuais: Player1 = nome do jogador, Position = colocação
    - Para duplas: Player1 + Player2 = parceiros, ambos recebem a mesma colocação
    """
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")

    tournament = await db.tournaments.find_one({"id": tournament_id}, {"_id": 0})
    if not tournament:
        raise HTTPException(status_code=404, detail="Torneio não encontrado")

    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content))

        players = await db.players.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
        players_dict = {p['name'].lower().strip(): p for p in players}

        config = await get_ranking_config()

        results_created = 0
        results_updated = 0
        errors = []

        def parse_category_from_sheet(sheet_name: str):
            """
            Extrai gender_category e class_category do nome da aba.
            Exemplos:
              '1ª Classe Masculina - 1ª Classe'  -> ('Masculina', '1ª')
              '3ª Classe Feminina - 3ª Classe '  -> ('Feminina', '3ª')
              'Duplas - Duplas'                   -> ('Mista', 'Duplas')
              'Principiante Masculina - Princi'   -> ('Masculina', 'Principiante')
            """
            name = sheet_name.strip()

            if 'dupla' in name.lower():
                return 'Mista', 'Duplas'

            if 'principiante' in name.lower():
                gender = 'Feminina' if 'feminina' in name.lower() or 'feminino' in name.lower() else 'Masculina'
                return normalize_gender(gender), 'Principiante'

            gender = 'Feminina' if 'feminina' in name.lower() or 'feminino' in name.lower() else 'Masculina'
            gender = normalize_gender(gender)

            # Detecta número ordinal: 1ª, 2ª, 3ª, etc.
            match = re.search(r'(\d+)[ªº°]', name)
            if match:
                num = match.group(1)
                class_map = {'1': '1ª', '2': '2ª', '3': '3ª', '4': '4ª', '5': '5ª', '6': '6ª'}
                class_category = class_map.get(num, f"{num}ª")
            else:
                class_category = 'Outra'

            return gender, class_category

        async def save_result(tournament_id, player, gender_category, class_category, placement, config):
            points = config.points_table.get(str(placement), 0.0)
            existing = await db.results.find_one({
                "tournament_id": tournament_id,
                "player_id": player['id'],
                "class_category": class_category,
                "gender_category": gender_category
            }, {"_id": 0})
            if existing:
                if placement < existing.get('placement', 999):
                    await db.results.update_one(
                        {"id": existing['id']},
                        {"$set": {"placement": placement, "points": points}}
                    )
                    return 'updated'
                return 'skipped'
            else:
                result = Result(
                    tournament_id=tournament_id,
                    player_id=player['id'],
                    player_name=player['name'],
                    class_category=class_category,
                    gender_category=gender_category,
                    placement=placement,
                    points=points
                )
                doc = result.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.results.insert_one(doc)
                return 'created'

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            gender_category, class_category = parse_category_from_sheet(sheet_name)
            is_doubles = (class_category == 'Duplas')

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[2]:  # Player1 é coluna C (índice 2)
                    continue

                try:
                    player1_name = str(row[2]).strip()
                    player2_name = str(row[3]).strip() if row[3] else None
                    position_val = row[4]  # Position (coluna E)

                    if not position_val:
                        continue

                    placement = int(float(position_val))

                    if is_doubles and player2_name:
                        # Duplas: salvar para os dois jogadores
                        for pname in [player1_name, player2_name]:
                            player = await get_or_create_player(pname, players_dict)
                            status = await save_result(tournament_id, player, gender_category, class_category, placement, config)
                            if status == 'created':
                                results_created += 1
                            elif status == 'updated':
                                results_updated += 1
                    else:
                        player = await get_or_create_player(player1_name, players_dict, gender_category if gender_category != 'Mista' else None)
                        status = await save_result(tournament_id, player, gender_category, class_category, placement, config)
                        if status == 'created':
                            results_created += 1
                        elif status == 'updated':
                            results_updated += 1

                except Exception as e:
                    errors.append(f"Aba '{sheet_name}' linha {row_idx}: {str(e)}")

        return {
            "message": f"{results_created} resultados importados, {results_updated} atualizados",
            "results_created": results_created,
            "results_updated": results_updated,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo: {str(e)}")

# Ranking Config Routes
@api_router.post("/admin/migrate-class-names")
async def migrate_class_names(current_user: User = Depends(get_current_user)):
    """
    Migração única: converte class_category de '1a','2a'... para '1ª','2ª'...
    e gender_category 'Mista' para registros de Duplas.
    Também atualiza main_class nos jogadores.
    """
    class_map = {'1a': '1ª', '2a': '2ª', '3a': '3ª', '4a': '4ª', '5a': '5ª', '6a': '6ª'}

    results_updated = 0
    matches_updated = 0
    players_updated = 0

    # Migra results
    for old, new in class_map.items():
        res = await db.results.update_many(
            {"class_category": old},
            {"$set": {"class_category": new}}
        )
        results_updated += res.modified_count

    # Migra matches (category field)
    for old, new in class_map.items():
        res = await db.matches.update_many(
            {"category": old},
            {"$set": {"category": new}}
        )
        matches_updated += res.modified_count

    # Migra players (main_class field)
    for old, new in class_map.items():
        res = await db.players.update_many(
            {"main_class": old},
            {"$set": {"main_class": new}}
        )
        players_updated += res.modified_count

    # Migra usuários existentes — define is_owner e is_approved
    owner_email = OWNER_EMAIL.lower()
    all_users = await db.users.find({}, {"_id": 0}).to_list(1000)
    for u in all_users:
        update = {}
        if "is_owner" not in u:
            update["is_owner"] = (u.get("username", "").lower() == owner_email)
        if "is_approved" not in u:
            # Owner e usuários já existentes antes da migração são aprovados automaticamente
            update["is_approved"] = True
        if update:
            await db.users.update_one({"id": u["id"]}, {"$set": update})

    # Normaliza qualquer valor antigo de gender_category
    for old_val, new_val in [('Masculino', 'Masculina'), ('Feminino', 'Feminina'), ('Misto', 'Mista')]:
        await db.results.update_many({"gender_category": old_val}, {"$set": {"gender_category": new_val}})
        await db.matches.update_many({"gender_category": old_val}, {"$set": {"gender_category": new_val}})
        await db.players.update_many({"gender": old_val}, {"$set": {"gender": new_val}})

    # Garante que resultados de Duplas tenham gender_category='Mista'
    duplas_fix = await db.results.update_many(
        {"class_category": "Duplas", "gender_category": {"$ne": "Mista"}},
        {"$set": {"gender_category": "Mista"}}
    )

    return {
        "message": "Migração concluída com sucesso",
        "results_updated": results_updated,
        "matches_updated": matches_updated,
        "players_updated": players_updated,
        "duplas_gender_fixed": duplas_fix.modified_count
    }



@api_router.get("/theme")
async def get_theme():
    """Retorna o tema atual salvo no banco"""
    config = await db.app_config.find_one({"key": "theme"}, {"_id": 0})
    if not config:
        return {"theme": "green"}
    return {"theme": config.get("value", "green")}

@api_router.put("/theme")
async def set_theme(theme: str, current_user: User = Depends(get_current_user)):
    """Salva o tema escolhido pelo admin"""
    valid_themes = ["green", "blue", "purple", "red", "orange", "silver"]
    if theme not in valid_themes:
        raise HTTPException(status_code=400, detail=f"Tema inválido. Opções: {valid_themes}")
    await db.app_config.update_one(
        {"key": "theme"},
        {"$set": {"key": "theme", "value": theme}},
        upsert=True
    )
    return {"theme": theme}

@api_router.get("/ranking-config", response_model=RankingConfig)
async def get_ranking_config():
    config = await db.ranking_config.find_one({}, {"_id": 0})
    if not config:
        # Create default config
        default_config = RankingConfig()
        doc = default_config.model_dump()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.ranking_config.insert_one(doc)
        return default_config
    
    if isinstance(config.get('updated_at'), str):
        config['updated_at'] = datetime.fromisoformat(config['updated_at'])
    return RankingConfig(**config)

@api_router.put("/ranking-config", response_model=RankingConfig)
async def update_ranking_config(config_data: RankingConfigUpdate, current_user: User = Depends(get_current_user)):
    config = await get_ranking_config()
    updated_data = config_data.model_dump()
    updated_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.ranking_config.update_one(
        {"id": config.id},
        {"$set": updated_data}
    )
    
    config_doc = await db.ranking_config.find_one({"id": config.id}, {"_id": 0})
    if isinstance(config_doc.get('updated_at'), str):
        config_doc['updated_at'] = datetime.fromisoformat(config_doc['updated_at'])
    return RankingConfig(**config_doc)

# Rankings Route
@api_router.get("/rankings")
async def get_rankings(class_category: Optional[str] = None, gender_category: Optional[str] = None):
    # ── 1. Config + results em paralelo ──────────────────────────────────────
    query = {}
    if class_category:
        query['class_category'] = class_category
    if gender_category:
        query['gender_category'] = gender_category

    config, results = await asyncio.gather(
        get_ranking_config(),
        db.results.find(query, {"_id": 0}).to_list(10000)
    )

    if not results:
        return []

    # ── 2. Agrupar resultados por jogador ────────────────────────────────────
    player_results = {}
    for result in results:
        pid = result['player_id']
        if pid not in player_results:
            player_results[pid] = {'player_id': pid, 'player_name': result['player_name'], 'results': []}
        player_results[pid]['results'].append(result['points'])

    player_ids = list(player_results.keys())

    # ── 3. Buscar TODOS os jogadores e TODAS as partidas em 2 queries únicas ─
    players_list, matches_list = await asyncio.gather(
        db.players.find({"id": {"$in": player_ids}}, {"_id": 0, "id": 1, "photo_url": 1, "is_federated": 1}).to_list(10000),
        db.matches.find(
            {"$or": [{"player1_id": {"$in": player_ids}}, {"player2_id": {"$in": player_ids}}]},
            {"_id": 0}
        ).sort("date", -1).to_list(10000)
    )

    # Indexar jogadores por id
    players_map = {p['id']: p for p in players_list}

    # Indexar último jogo por player_id (lista já está ordenada por date desc)
    last_match_map = {}
    for match in matches_list:
        for pid_key in [match['player1_id'], match['player2_id']]:
            if pid_key not in last_match_map:
                last_match_map[pid_key] = match

    # ── 4. Calcular pontos e montar resposta ─────────────────────────────────
    rankings = []
    for pid, data in player_results.items():
        points_list = sorted(data['results'], reverse=True)
        player = players_map.get(pid, {})
        is_federated = player.get('is_federated', True)

        if config.formula == "sum_all":
            total_points = sum(points_list) if is_federated else 0.0
        elif config.formula == "top_n":
            total_points = sum(points_list[:config.top_n_count]) if is_federated else 0.0
        elif config.formula == "decay":
            total_points = sum([p * (0.9 ** i) for i, p in enumerate(points_list)]) if is_federated else 0.0
        else:
            total_points = sum(points_list) if is_federated else 0.0

        last_match = None
        last_match_doc = last_match_map.get(pid)
        if last_match_doc:
            is_player1 = last_match_doc['player1_id'] == pid
            set_result = calculate_set_result(last_match_doc['score'])
            last_match = {
                "opponent_name": last_match_doc['player2_name'] if is_player1 else last_match_doc['player1_name'],
                "score": last_match_doc['score'],
                "score_formatted": f"{', '.join(last_match_doc['score'])} ({set_result})",
                "set_result": set_result,
                "result": "Win" if last_match_doc['winner_id'] == pid else "Loss",
                "tournament_name": last_match_doc.get('tournament_name', ''),
                "date": last_match_doc['date']
            }

        rankings.append({
            'player_id': pid,
            'player_name': data['player_name'],
            'photo_url': player.get('photo_url'),
            'total_points': round(total_points, 2),
            'results_count': len(points_list),
            'is_federated': is_federated,
            'last_match': last_match
        })

    rankings.sort(key=lambda x: x['total_points'], reverse=True)
    for i, r in enumerate(rankings, 1):
        r['rank'] = i

    return rankings

# Import from image
@api_router.post("/import-from-image")
async def import_from_image(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    try:
        # Read image
        contents = await file.read()
        base64_image = base64.b64encode(contents).decode('utf-8')
        
        # 
        prompt = """Analise esta imagem de uma chave/bracket de torneio de squash.
        
Extraia TODOS os jogadores e suas colocações finais. Para cada jogador, identifique:
- player_name: nome completo
- placement: colocação numérica (1=campeão, 2=vice, 3/4=semifinalistas, 5-8=perdedores de quartas)
- category: "Masculina" ou "Feminina"

Retorne os dados no formato JSON:
{
  "tournament_name": "Nome do Torneio",
  "results": [
    {"player_name": "Nome", "placement": 1, "category": "Masculina"}
  ]
}"""

        genai.configure(api_key=os.environ.get('GEMINI_API_KEY', ''))
        model = genai.GenerativeModel('gemini-1.5-flash')
        image_part = {"mime_type": file.content_type, "data": contents}
        response = await asyncio.to_thread(model.generate_content, [prompt, image_part])
        response_text = response.text
        
        # Parse JSON from response
        import json
        
        # Try to extract JSON from markdown code blocks or raw text
        json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', response_text, re.DOTALL)
        if json_match:
            json_str = json_match.group(1)
        else:
            # Try to find JSON object in the response
            json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
            if json_match:
                json_str = json_match.group(0)
            else:
                json_str = response_text
        
        result = json.loads(json_str)
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing image: {str(e)}")

# Match Routes
@api_router.get("/matches", response_model=List[Match])
async def get_matches(player_id: Optional[str] = None, tournament_id: Optional[str] = None):
    query = {}
    if player_id:
        query["$or"] = [{"player1_id": player_id}, {"player2_id": player_id}]
    if tournament_id:
        query["tournament_id"] = tournament_id
    
    matches = await db.matches.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    for match in matches:
        if isinstance(match.get('date'), str):
            match['date'] = datetime.fromisoformat(match['date'])
        if isinstance(match.get('created_at'), str):
            match['created_at'] = datetime.fromisoformat(match['created_at'])
    return matches

@api_router.post("/matches", response_model=Match)
async def create_match(match_data: MatchCreate, current_user: User = Depends(get_current_user)):
    # Get tournament name
    tournament = await db.tournaments.find_one({"id": match_data.tournament_id}, {"_id": 0})
    if not tournament:
        raise HTTPException(status_code=404, detail="Tournament not found")
    
    # Get player names
    player1 = await db.players.find_one({"id": match_data.player1_id}, {"_id": 0})
    player2 = await db.players.find_one({"id": match_data.player2_id}, {"_id": 0})
    
    if not player1 or not player2:
        raise HTTPException(status_code=404, detail="Player not found")
    
    match = Match(
        **match_data.model_dump(),
        tournament_name=tournament['name'],
        player1_name=player1['name'],
        player2_name=player2['name']
    )
    
    doc = match.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.matches.insert_one(doc)
    
    # ============= AUTO-CREATE/UPDATE RESULTS =============
    # Calculate placements based on round
    winner_placement = calculate_placement_from_round(match.round, True)
    loser_placement = calculate_placement_from_round(match.round, False)
    
    # Get ranking config for points
    config = await get_ranking_config()
    winner_points = config.points_table.get(str(winner_placement), 0.0)
    loser_points = config.points_table.get(str(loser_placement), 0.0)
    
    # Determine winner and loser
    winner_id = match.winner_id
    loser_id = match.player2_id if winner_id == match.player1_id else match.player1_id
    winner_name = match.player1_name if winner_id == match.player1_id else match.player2_name
    loser_name = match.player2_name if winner_id == match.player1_id else match.player1_name
    
    # Check if results already exist for these players in this tournament and category
    # Update or create winner result
    existing_winner_result = await db.results.find_one({
        "tournament_id": match.tournament_id,
        "player_id": winner_id,
        "class_category": match.category,
        "gender_category": "Masculina"  # Default, could be enhanced
    }, {"_id": 0})
    
    if existing_winner_result:
        # Update if new placement is better (lower number = better)
        if winner_placement < existing_winner_result.get('placement', 999):
            await db.results.update_one(
                {"id": existing_winner_result['id']},
                {"$set": {
                    "placement": winner_placement,
                    "points": winner_points
                }}
            )
    else:
        # Create new result
        winner_result = Result(
            tournament_id=match.tournament_id,
            player_id=winner_id,
            player_name=winner_name,
            class_category=match.category,
            gender_category="Masculina",  # Default
            placement=winner_placement,
            points=winner_points
        )
        winner_doc = winner_result.model_dump()
        winner_doc['created_at'] = winner_doc['created_at'].isoformat()
        await db.results.insert_one(winner_doc)
    
    # Update or create loser result
    existing_loser_result = await db.results.find_one({
        "tournament_id": match.tournament_id,
        "player_id": loser_id,
        "class_category": match.category,
        "gender_category": "Masculina"
    }, {"_id": 0})
    
    if existing_loser_result:
        # Update if new placement is better
        if loser_placement < existing_loser_result.get('placement', 999):
            await db.results.update_one(
                {"id": existing_loser_result['id']},
                {"$set": {
                    "placement": loser_placement,
                    "points": loser_points
                }}
            )
    else:
        # Create new result
        loser_result = Result(
            tournament_id=match.tournament_id,
            player_id=loser_id,
            player_name=loser_name,
            class_category=match.category,
            gender_category="Masculina",
            placement=loser_placement,
            points=loser_points
        )
        loser_doc = loser_result.model_dump()
        loser_doc['created_at'] = loser_doc['created_at'].isoformat()
        await db.results.insert_one(loser_doc)
    
    return match

@api_router.delete("/matches/{match_id}")
async def delete_match(match_id: str, current_user: User = Depends(get_current_user)):
    result = await db.matches.delete_one({"id": match_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Match not found")
    return {"message": "Match deleted"}

@api_router.get("/matches/template")
async def download_matches_template():
    """Generate and download Excel template for matches import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"
    
    # Headers
    headers = ["Tournament", "Category", "Round", "Player1", "Player2", "Score", "Winner", "Date"]
    ws.append(headers)
    
    # Example row
    example = ["Copa PR", "1ª", "Final", "João Silva", "Pedro Lima", "11-7 8-11 11-6", "João Silva", "2025-01-15"]
    ws.append(example)
    
    # Style headers (bold)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_partidas.xlsx"}
    )

# Import matches from Excel
@api_router.post("/import-matches-excel")
async def import_matches_excel(
    tournament_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Importa partidas de arquivo Excel do Tournament Planner.
    Formato: múltiplas abas por dia (ex: 'sexta-feira 27 fev 2026')
    Cada aba tem 3 linhas de cabeçalho, depois a linha de colunas na linha 4:
    Time | Event | Nr | Court | Location | Round | Team 1 | Team 2 | Score | Duration | Referee | Marker | Start | Finish
    - Team 1 e Team 2 podem ser 'Jogador A+Jogador B' para duplas
    - Score: ex '11-9 11-4 11-5 ' (sets separados por espaço)
    - O vencedor é determinado contando sets ganhos por cada lado
    """
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")

    tournament = await db.tournaments.find_one({"id": tournament_id}, {"_id": 0})
    if not tournament:
        raise HTTPException(status_code=404, detail="Torneio não encontrado")

    try:
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))

        players = await db.players.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
        players_dict = {p['name'].lower().strip(): p for p in players}

        matches_created = 0
        matches_skipped = 0
        errors = []

        def determine_winner(score_str: str, player1_name: str, player2_name: str):
            """
            Conta sets do score '11-9 11-4 5-11 ' e retorna nome do vencedor.
            Cada set: lado esquerdo = Team 1, direito = Team 2.
            """
            if not score_str or score_str.strip().lower() in ('w.o.', 'wo', 'walkover', ''):
                return player1_name  # w.o. = Team 1 vence por W.O.
            sets = [s.strip() for s in score_str.strip().split() if s.strip()]
            p1_sets = 0
            p2_sets = 0
            for s in sets:
                parts = s.split('-')
                if len(parts) == 2:
                    try:
                        s1, s2 = int(parts[0]), int(parts[1])
                        if s1 > s2:
                            p1_sets += 1
                        else:
                            p2_sets += 1
                    except ValueError:
                        pass
            return player1_name if p1_sets >= p2_sets else player2_name

        def resolve_player(name_raw: str):
            """
            Limpa sufixos de seed como '[1]', '[3/4]' do nome antes de buscar.
            """
            name_clean = re.sub(r'\s*\[.*?\]', '', name_raw).strip()
            return players_dict.get(name_clean.lower()), name_clean

        def parse_event_category(event_name: str):
            """
            Converte nome do evento para class_category e gender_category.
            '4ª Classe Masculina' -> ('4ª', 'Masculina')
            'Duplas'              -> ('Duplas', 'Mista')
            'Principiante Masculina' -> ('Principiante', 'Masculina')
            """
            if 'dupla' in event_name.lower():
                return 'Duplas', 'Mista'
            if 'principiante' in event_name.lower():
                gender = 'Feminina' if 'feminina' in event_name.lower() or 'feminino' in event_name.lower() else 'Masculina'
                return 'Principiante', gender
            gender = 'Feminina' if 'feminina' in event_name.lower() or 'feminino' in event_name.lower() else 'Masculina'
            m = re.search(r'(\d+)[ªº°]', event_name)
            if m:
                num = m.group(1)
                class_map = {'1': '1ª', '2': '2ª', '3': '3ª', '4': '4ª', '5': '5ª', '6': '6ª'}
                class_category = class_map.get(num, f"{num}ª")
            else:
                class_category = 'Outra'
            return class_category, gender

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Cada aba tem 3 linhas de header do Tournament Planner, depois a linha de colunas (row 4)
            # Dados começam na linha 5
            for row_idx, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
                if not row or not row[0]:
                    continue

                try:
                    time_val   = row[0]   # 'sex 27/02/2026 19:00'
                    event_name = str(row[1]).strip() if row[1] else ""
                    round_name = str(row[5]).strip() if row[5] else ""
                    team1_raw  = str(row[6]).strip() if row[6] else ""
                    team2_raw  = str(row[7]).strip() if row[7] else ""
                    score_str  = str(row[8]).strip() if row[8] else ""

                    if not event_name or not team1_raw or not team2_raw:
                        continue

                    class_category, gender_category = parse_event_category(event_name)
                    is_doubles = class_category == 'Duplas'

                    # Parse date from time_val string like 'sex 27/02/2026 19:00'
                    match_date = None
                    if isinstance(time_val, datetime):
                        match_date = time_val
                    else:
                        time_str = str(time_val).strip()
                        date_match = re.search(r'(\d{2}/\d{2}/\d{4})', time_str)
                        if date_match:
                            match_date = datetime.strptime(date_match.group(1), '%d/%m/%Y')
                        else:
                            match_date = datetime.now(timezone.utc)

                    if is_doubles:
                        # Duplas: 'Jogador A+Jogador B' vs 'Jogador C+Jogador D'
                        def split_doubles(team_raw):
                            name_clean = re.sub(r'\s*\[.*?\]', '', team_raw).strip()
                            if '+' in name_clean:
                                parts = [p.strip() for p in name_clean.split('+')]
                                return parts[0], parts[1]
                            return name_clean, None

                        p1a_name, p1b_name = split_doubles(team1_raw)
                        p2a_name, p2b_name = split_doubles(team2_raw)

                        p1a = await get_or_create_player(p1a_name, players_dict) if p1a_name else None
                        p1b = await get_or_create_player(p1b_name, players_dict) if p1b_name else None
                        p2a = await get_or_create_player(p2a_name, players_dict) if p2a_name else None
                        p2b = await get_or_create_player(p2b_name, players_dict) if p2b_name else None

                        winner_team = determine_winner(score_str, p1a_name, p2a_name)
                        winner_id = (p1a or p1b)['id'] if winner_team == p1a_name else (p2a or p2b)['id']
                        winner_name = winner_team

                        # Registra a partida com player1 = primeiro da dupla 1
                        p1 = p1a or p1b
                        p2 = p2a or p2b
                        player1_name_full = f"{p1a_name}+{p1b_name}" if p1b_name else p1a_name
                        player2_name_full = f"{p2a_name}+{p2b_name}" if p2b_name else p2a_name
                    else:
                        p1, p1_clean = resolve_player(team1_raw)
                        p2, p2_clean = resolve_player(team2_raw)

                        if not p1:
                            p1 = await get_or_create_player(p1_clean, players_dict)
                        if not p2:
                            p2 = await get_or_create_player(p2_clean, players_dict)

                        winner_name = determine_winner(score_str, p1['name'], p2['name'])
                        winner_id = p1['id'] if winner_name == p1['name'] else p2['id']
                        player1_name_full = p1['name']
                        player2_name_full = p2['name']

                    score_list = [s for s in score_str.split() if s.strip()] if score_str else []

                    match = Match(
                        tournament_id=tournament['id'],
                        tournament_name=tournament['name'],
                        category=class_category,
                        player1_id=p1['id'],
                        player1_name=player1_name_full,
                        player2_id=p2['id'],
                        player2_name=player2_name_full,
                        winner_id=winner_id,
                        score=score_list,
                        round=round_name,
                        date=match_date
                    )
                    doc = match.model_dump()
                    doc['date'] = doc['date'].isoformat() if isinstance(doc['date'], datetime) else doc['date']
                    doc['created_at'] = doc['created_at'].isoformat()

                    # Verifica duplicidade: mesma partida = mesmo torneio + mesmos jogadores + mesma rodada
                    existing_match = await db.matches.find_one({
                        "tournament_id": tournament['id'],
                        "player1_id": p1['id'],
                        "player2_id": p2['id'],
                        "round": round_name
                    })
                    # Tenta também na ordem inversa (p1 e p2 podem estar trocados)
                    if not existing_match:
                        existing_match = await db.matches.find_one({
                            "tournament_id": tournament['id'],
                            "player1_id": p2['id'],
                            "player2_id": p1['id'],
                            "round": round_name
                        })

                    if existing_match:
                        matches_skipped += 1
                    else:
                        await db.matches.insert_one(doc)
                        matches_created += 1

                except Exception as e:
                    errors.append(f"Aba '{sheet_name}' linha {row_idx}: {str(e)}")

        return {
            "matches_created": matches_created,
            "matches_skipped": matches_skipped,
            "errors": errors
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar Excel: {str(e)}")

# ============= SUBSCRIPTION ENDPOINTS =============

# Register federation with trial or paid plan
@api_router.post("/register-federation")
async def register_federation(federation_data: FederationCreate):
    # Check if username already exists
    existing_user = await db.users.find_one({"username": federation_data.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    hashed_password = pwd_context.hash(federation_data.password)
    user = User(
        username=federation_data.email,
        hashed_password=hashed_password,
        federation_name=federation_data.federation_name,
        is_admin=True
    )
    
    # Create subscription
    if federation_data.start_trial:
        # 1 day trial
        subscription = Subscription(
            user_id=user.id,
            plan_type=federation_data.plan_type,
            status="active",
            is_trial=True,
            start_date=datetime.now(timezone.utc),
            end_date=datetime.now(timezone.utc) + timedelta(days=1),
            next_billing_date=datetime.now(timezone.utc) + timedelta(days=1)
        )
    else:
        # Paid plan - will be activated after payment
        plan = SUBSCRIPTION_PLANS[federation_data.plan_type]
        subscription = Subscription(
            user_id=user.id,
            plan_type=federation_data.plan_type,
            status="pending",  # Pending payment
            is_trial=False,
            start_date=datetime.now(timezone.utc),
            end_date=datetime.now(timezone.utc) + timedelta(days=plan['duration_days']),
            next_billing_date=datetime.now(timezone.utc) + timedelta(days=plan['duration_days'])
        )
    
    # Update user with subscription_id
    user.subscription_id = subscription.id
    
    # Save to database
    user_doc = user.model_dump()
    user_doc['created_at'] = user_doc['created_at'].isoformat()
    await db.users.insert_one(user_doc)
    
    subscription_doc = subscription.model_dump()
    subscription_doc['start_date'] = subscription_doc['start_date'].isoformat()
    subscription_doc['end_date'] = subscription_doc['end_date'].isoformat()
    if subscription_doc['next_billing_date']:
        subscription_doc['next_billing_date'] = subscription_doc['next_billing_date'].isoformat()
    subscription_doc['created_at'] = subscription_doc['created_at'].isoformat()
    await db.subscriptions.insert_one(subscription_doc)
    
    # Send email notification
    if federation_data.start_trial:
        try:
            params = {
                "from": "noreply@fsp.com.br",
                "to": [federation_data.email],
                "subject": "Trial ativado - SquashRank Pro",
                "html": f"""
                    <h2>Bem-vindo ao SquashRank Pro!</h2>
                    <p>Olá {federation_data.federation_name},</p>
                    <p>Seu trial de 1 dia foi ativado com sucesso!</p>
                    <p>Você tem até <strong>{subscription.end_date.strftime('%d/%m/%Y às %H:%M')}</strong> para explorar todas as funcionalidades.</p>
                    <p>Após o trial, você precisará escolher um plano para continuar usando o sistema.</p>
                    <p>Obrigado por escolher SquashRank Pro!</p>
                """
            }
            resend.Emails.send(params)
        except Exception as e:
            logger.error(f"Error sending trial email: {e}")
    
    return {
        "message": "Federation registered successfully",
        "user_id": user.id,
        "subscription_id": subscription.id,
        "is_trial": subscription.is_trial,
        "end_date": subscription.end_date.isoformat()
    }

# Create Stripe checkout session
@api_router.post("/subscriptions/create-checkout")
async def create_checkout_session(
    plan_type: str,
    origin_url: str,
    current_user: User = Depends(get_current_user)
):
    """Create Stripe checkout session for subscription payment"""
    if plan_type not in SUBSCRIPTION_PLANS:
        raise HTTPException(status_code=400, detail="Invalid plan type")
    
    plan = SUBSCRIPTION_PLANS[plan_type]
    
    # Initialize Stripe
    stripe_api_key = os.getenv('STRIPE_API_KEY')
    webhook_url = f"{origin_url}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=stripe_api_key, webhook_url=webhook_url)
    
    # Create checkout session
    success_url = f"{origin_url}/admin/dashboard?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin_url}/admin/dashboard"
    
    checkout_request = CheckoutSessionRequest(
        amount=plan['price'],
        currency=plan['currency'],
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={
            "user_id": current_user.id,
            "plan_type": plan_type,
            "federation_name": current_user.federation_name or ""
        }
    )
    
    session: CheckoutSessionResponse = await stripe_checkout.create_checkout_session(checkout_request)
    
    # Create payment transaction record
    transaction_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user.id,
        "session_id": session.session_id,
        "amount": plan['price'],
        "currency": plan['currency'],
        "payment_status": "pending",
        "metadata": checkout_request.metadata,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.payment_transactions.insert_one(transaction_doc)
    
    return {"url": session.url, "session_id": session.session_id}

# Get checkout status
@api_router.get("/subscriptions/status/{session_id}")
async def get_checkout_status(
    session_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get checkout session status and update subscription if paid"""
    # Get transaction
    transaction = await db.payment_transactions.find_one(
        {"session_id": session_id, "user_id": current_user.id},
        {"_id": 0}
    )
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # If already paid, return cached status
    if transaction['payment_status'] == 'paid':
        return {
            "payment_status": "paid",
            "message": "Payment already processed"
        }
    
    # Check with Stripe
    stripe_api_key = os.getenv('STRIPE_API_KEY')
    stripe_checkout = StripeCheckout(api_key=stripe_api_key, webhook_url="")
    
    try:
        checkout_status: CheckoutStatusResponse = await stripe_checkout.get_checkout_status(session_id)
        
        # Update transaction
        await db.payment_transactions.update_one(
            {"session_id": session_id},
            {"$set": {
                "payment_status": checkout_status.payment_status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # If paid and not yet processed, activate subscription
        if checkout_status.payment_status == 'paid' and transaction['payment_status'] != 'paid':
            metadata = checkout_status.metadata
            user_id = metadata.get('user_id')
            plan_type = metadata.get('plan_type')
            
            if user_id and plan_type:
                # Get user's subscription
                subscription = await db.subscriptions.find_one(
                    {"user_id": user_id},
                    {"_id": 0}
                )
                
                if subscription:
                    plan = SUBSCRIPTION_PLANS[plan_type]
                    
                    # Update subscription to active
                    await db.subscriptions.update_one(
                        {"user_id": user_id},
                        {"$set": {
                            "status": "active",
                            "is_trial": False,
                            "plan_type": plan_type,
                            "start_date": datetime.now(timezone.utc).isoformat(),
                            "end_date": (datetime.now(timezone.utc) + timedelta(days=plan['duration_days'])).isoformat(),
                            "next_billing_date": (datetime.now(timezone.utc) + timedelta(days=plan['duration_days'])).isoformat()
                        }}
                    )
                    
                    # Send confirmation email
                    try:
                        user = await db.users.find_one({"id": user_id}, {"_id": 0})
                        if user:
                            params = {
                                "from": "noreply@fsp.com.br",
                                "to": [user['username']],
                                "subject": "Pagamento confirmado - SquashRank Pro",
                                "html": f"""
                                    <h2>Pagamento Confirmado!</h2>
                                    <p>Olá {user.get('federation_name', 'Federação')},</p>
                                    <p>Seu pagamento foi confirmado com sucesso!</p>
                                    <p><strong>Plano:</strong> {plan['name']}</p>
                                    <p><strong>Valor:</strong> R$ {plan['price']:.2f}</p>
                                    <p>Sua assinatura está ativa e válida até {(datetime.now(timezone.utc) + timedelta(days=plan['duration_days'])).strftime('%d/%m/%Y')}.</p>
                                    <p>Obrigado por escolher SquashRank Pro!</p>
                                """
                            }
                            resend.Emails.send(params)
                    except Exception as e:
                        logger.error(f"Error sending payment confirmation email: {e}")
        
        return {
            "payment_status": checkout_status.payment_status,
            "status": checkout_status.status,
            "amount_total": checkout_status.amount_total,
            "currency": checkout_status.currency
        }
        
    except Exception as e:
        logger.error(f"Error checking checkout status: {e}")
        raise HTTPException(status_code=500, detail="Error checking payment status")

# Get my subscription
@api_router.get("/subscriptions/my-subscription")
async def get_my_subscription(current_user: User = Depends(get_current_user)):
    """Get current user's subscription status"""
    subscription = await db.subscriptions.find_one(
        {"user_id": current_user.id},
        {"_id": 0}
    )
    
    if not subscription:
        return {
            "has_subscription": False,
            "message": "No subscription found"
        }
    
    # Check if expired
    end_date = datetime.fromisoformat(subscription['end_date'].replace('Z', '+00:00'))
    is_expired = datetime.now(timezone.utc) >= end_date
    
    if is_expired and subscription['status'] == 'active':
        # Mark as expired
        await db.subscriptions.update_one(
            {"user_id": current_user.id},
            {"$set": {"status": "expired"}}
        )
        subscription['status'] = 'expired'
    
    # Calculate days remaining
    days_remaining = (end_date - datetime.now(timezone.utc)).days
    
    return {
        "has_subscription": True,
        "subscription": subscription,
        "days_remaining": max(0, days_remaining),
        "is_active": subscription['status'] == 'active' and not is_expired
    }

# Stripe webhook
@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """Handle Stripe webhook events"""
    try:
        body = await request.body()
        signature = request.headers.get("Stripe-Signature")
        
        stripe_api_key = os.getenv('STRIPE_API_KEY')
        stripe_checkout = StripeCheckout(api_key=stripe_api_key, webhook_url="")
        
        webhook_response = await stripe_checkout.handle_webhook(body, signature)
        
        logger.info(f"Webhook received: {webhook_response.event_type}")
        
        # Handle different event types
        if webhook_response.event_type == "checkout.session.completed":
            session_id = webhook_response.session_id
            metadata = webhook_response.metadata
            
            # Update transaction
            await db.payment_transactions.update_one(
                {"session_id": session_id},
                {"$set": {
                    "payment_status": webhook_response.payment_status,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Activate subscription if paid
            if webhook_response.payment_status == 'paid':
                user_id = metadata.get('user_id')
                plan_type = metadata.get('plan_type')
                
                if user_id and plan_type:
                    plan = SUBSCRIPTION_PLANS.get(plan_type)
                    if plan:
                        await db.subscriptions.update_one(
                            {"user_id": user_id},
                            {"$set": {
                                "status": "active",
                                "is_trial": False,
                                "plan_type": plan_type,
                                "start_date": datetime.now(timezone.utc).isoformat(),
                                "end_date": (datetime.now(timezone.utc) + timedelta(days=plan['duration_days'])).isoformat()
                            }}
                        )
        
        return {"status": "success"}
        
    except Exception as e:
        logger.error(f"Webhook error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

# Dependency to check active subscription
async def check_active_subscription(current_user: User = Depends(get_current_user)):
    """Middleware to check if user has active subscription"""
    subscription = await db.subscriptions.find_one(
        {"user_id": current_user.id},
        {"_id": 0}
    )
    
    if not subscription:
        raise HTTPException(
            status_code=403,
            detail="No subscription found. Please subscribe to access this feature."
        )
    
    # Check if expired
    end_date = datetime.fromisoformat(subscription['end_date'].replace('Z', '+00:00'))
    is_expired = datetime.now(timezone.utc) >= end_date
    
    if is_expired or subscription['status'] not in ['active']:
        raise HTTPException(
            status_code=403,
            detail="Your subscription has expired. Please renew to continue using the system."
        )
    
    return current_user

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def create_indexes():
    """Cria índices no MongoDB para acelerar as queries críticas."""
    try:
        # Índices na coleção results — usados no endpoint /rankings
        await db.results.create_index([("class_category", 1), ("gender_category", 1)])
        await db.results.create_index([("player_id", 1)])

        # Índices na coleção players
        await db.players.create_index([("id", 1)], unique=True, sparse=True)
        await db.players.create_index([("name", 1)])

        # Índices na coleção matches — busca de último jogo por player
        await db.matches.create_index([("player1_id", 1), ("date", -1)])
        await db.matches.create_index([("player2_id", 1), ("date", -1)])
        await db.matches.create_index([("date", -1)])

        # Índices na coleção tournaments
        await db.tournaments.create_index([("date", -1)])

        logger.info("Índices do MongoDB criados com sucesso.")
    except Exception as e:
        logger.warning(f"Erro ao criar índices: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
